import streamlit as st
import pandas as pd
import zipfile
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def copy_style(source_cell, dest_cell):
    # Kopya styles from original sheet
    dest_cell.font = openpyxl.styles.Font(name=source_cell.font.name,
                                         size=source_cell.font.size,
                                         bold=source_cell.font.bold,
                                         color=source_cell.font.color)
    dest_cell.fill = openpyxl.styles.PatternFill(start_color=source_cell.fill.start_color,
                                                end_color=source_cell.fill.end_color,
                                                fill_type=source_cell.fill.fill_type)
    dest_cell.border = openpyxl.styles.Border(left=source_cell.border.left,
                                            right=source_cell.border.right,
                                            top=source_cell.border.top,
                                            bottom=source_cell.border.bottom)
    dest_cell.alignment = openpyxl.styles.Alignment(horizontal=source_cell.alignment.horizontal,
                                                  vertical=source_cell.alignment.vertical,
                                                  wrap_text=source_cell.alignment.wrap_text)


def split_and_save_files(df, column_to_filter, sheet_name, num_splits, progress_bar):
    unique_values = df[column_to_filter].dropna().unique()
    
    # Create dict to store the splits
    split_data = {i: [] for i in range(num_splits)}
    
    value_to_split = {}
    
    for i, value in enumerate(unique_values):
        split_index = i % num_splits
        split_data[split_index].append(value)
        value_to_split[value] = split_index
    
    os.makedirs("split_files", exist_ok=True)
    
    for split_index in range(num_splits):
        # Create new wb
        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        output_sheet.title = sheet_name  # Set the sheet name provided by the user

        split_dfs = []

        for _, row in df.iterrows():
            value = row[column_to_filter]
            if not pd.isna(value) and value_to_split[value] == split_index:
                split_dfs.append(row)

        split_df = pd.DataFrame(split_dfs)

        for r_idx, row in enumerate(dataframe_to_rows(split_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = output_sheet.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:  # Header row for style
                    source_cell = details_sheet.cell(row=r_idx, column=c_idx)
                    copy_style(source_cell, cell)  # Copy style from Details sheet
                    cell.number_format = source_cell.number_format  # Copy number format

        # Save the split sheet
        split_file_name = f"{filename}_split_{split_index}.xlsx"
        output_wb.save(split_file_name)
        progress_bar.progress((split_index + 1) / num_splits)  # Update the progress bar

    # Create a ZIP file to compress the sheets for dl
    zip_file_name = "split_files.zip"
    with zipfile.ZipFile(zip_file_name, 'w') as zipf:
        for split_index in range(num_splits):
            split_file_name = f"{filename}_split_{split_index}.xlsx"
            zipf.write(split_file_name)
            os.remove(split_file_name)  # Remove the individual split file

    return zip_file_name

st.title("Excel Data Splitter")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    filename = os.path.splitext(uploaded_file.name)[0]
    
    try:
        xls = pd.ExcelFile(uploaded_file)
        sheet_names = xls.sheet_names
        sheet_name = st.selectbox("Select the sheet to split:", sheet_names, index=0)
        # Input for the number of rows to skip
        rows_to_skip = st.number_input("Enter the number of rows to skip", min_value=0, max_value=None, value=0)

        details_df = xls.parse(sheet_name)
        details_sheet = xls.book[sheet_name]  # Get the user-specified sheet
    except Exception as e:
        st.error(f"Error reading the Excel file: {e}")
    else:
        st.write("Uploaded Excel Data:")
        st.dataframe(details_df)

        # Get the name of the column to filter
        column_to_filter = st.text_input("Enter the column name to base the splitting:")

        # Input for the number of splits
        num_splits = st.number_input("Enter the number of splits", min_value=1, max_value=None, value=3)

        if st.button("Split Details"):
            # Create a progress bar
            progress_bar = st.progress(0)
            # Perform the split and save split files
            zip_file_name = split_and_save_files(details_df.iloc[rows_to_skip:], column_to_filter, sheet_name, num_splits, progress_bar)

            # Provide a download link for the ZIP file
            with open(zip_file_name, "rb") as file:
                st.download_button(
                    label="Download Splitted Files",
                    data=file,
                    key=zip_file_name,
                    file_name=zip_file_name,
                )